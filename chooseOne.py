import glob, os #поиск ФАЙЛА АВТОМАТИЧЕСКИ
from openpyxl import load_workbook
class example3:
    def choose1(mainFolder):
        correct = 0  # пока пользователь не выполнит корректно операцию, программа дальше не пойдет
        while correct == 0:
            print(
                "Выбрать первую в списке таблицу из данной дериктории(1), или ввести название файла если их несколько в папке(2)?")
            way = input()  # выбор пользователя
            if way == "1":
                print("открывааем первый найденный файл xlsx")
                try:
                    os.chdir(mainFolder)  # формируем список файлов xlsx в папке и берем первый
                    nameFile = str(glob.glob("*.xlsx")[0])
                    book = load_workbook(filename=f"{mainFolder}/{nameFile}")
                    correct = 1
                except:
                    print("допустимый формат не обнаружен")
            elif way == "2":
                print("Введите название одного из файлов в папке(с расширением пожалуйста)")
                try:
                    nameFile = input()
                    book = load_workbook(filename=f"{mainFolder}/{nameFile}")
                    correct = 1
                except:
                    print("нет такого файла, давай по новой")
            else:
                print("были введены некорректные параметры. Попробуйте начать заново")
        return nameFile, book